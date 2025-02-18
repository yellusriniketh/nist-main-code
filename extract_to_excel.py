import os
import re
import openpyxl

def extract_data_to_excel(directory, output_excel_filename="oxygen.xlsx"):
    """
    Extracts data from .txt files in a directory and saves it to an Excel file,
    with each .txt file's data in a separate sheet.  Sheet names are derived
    from pressure and temperature values in the filename.
    """

    workbook = openpyxl.Workbook()
    # Remove the default sheet
    del workbook['Sheet']

    for filename in os.listdir(directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(directory, filename)

            # Extract pressure and temperature from filename using regex
            match = re.search(r"P-(\d+)_T-(\d+)-(\d+)", filename)
            if match:
                pressure = match.group(1)
                temp_low = match.group(2)
                temp_high = match.group(3)
                sheet_name = f"P-{pressure}_T-{temp_low}-{temp_high}"
            else:
                sheet_name = filename.replace(".txt", "")  # Default sheet name

            # Create a new worksheet
            worksheet = workbook.create_sheet(title=sheet_name)

            with open(filepath, 'r') as file:
                data_lines = file.readlines()

                # Write data to the worksheet
                for row_index, line in enumerate(data_lines, start=1):
                    values = line.strip().split()  # Split line into values
                    for col_index, value in enumerate(values, start=1):
                        try:
                            worksheet.cell(row=row_index, column=col_index, value=float(value))
                        except ValueError:
                            worksheet.cell(row=row_index, column=col_index, value=value)


    # Save the workbook
    workbook.save(output_excel_filename)
    print(f"Data extracted and saved to {output_excel_filename}")

if __name__ == "__main__":
    # Example usage:
    directory_path = "/workspaces/nist-2"  # Replace with the directory containing your .txt files
    output_file_name = "nist_data.xlsx"
    extract_data_to_excel(directory_path, output_file_name)
